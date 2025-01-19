import asyncio
import logging
import sys
from os import getenv

from aiogram import Bot, Dispatcher, html
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart
from aiogram.types import Message
from aiogram import types
from aiogram.filters.command import Command
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv

load_dotenv()

TOKEN = getenv("BOT_TOKEN")
EXCEL_FILE_FB = getenv("EXCEL_FILE_FB")
EXCEL_FILE_Q = getenv("EXCEL_FILE_Q")

dp = Dispatcher()

def save_feedback(username, message):
    try:
        try:
            workbook = load_workbook(EXCEL_FILE_FB)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Username", "Message", "Date"])
        sheet.append([username, message, "=NOW()"])
        workbook.save(EXCEL_FILE_FB)
    except Exception as e:
        print(f"Ошибка при записи в Excel: {e}")

def save_question(username, message):
    try:
        try:
            workbook = load_workbook(EXCEL_FILE_Q)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Username", "Message", "Date"])
        sheet.append([username, message, "=NOW()"])
        workbook.save(EXCEL_FILE_Q)
    except Exception as e:
        print(f"Ошибка при записи в Excel: {e}")


@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    kb = [
        [types.KeyboardButton(text="Написать отзыв")],
        [types.KeyboardButton(text="Задать вопрос")]
    ]
    keyboard = types.ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)
    await message.answer_sticker(getenv("HELLO_STICKER"))
    await message.answer(
        f"Здравствуйте, {html.bold(message.from_user.full_name)}!",
        reply_markup=keyboard
    )

feedback_users = set()
questions = set()
@dp.message()
async def handle_message(message: Message):
    if message.text == "Написать отзыв":
        feedback_users.add(message.from_user.id)
        await message.answer("Напишите свой отзыв")

    elif message.from_user.id in feedback_users:
        save_feedback(message.from_user.username or "No username", message.text)
        feedback_users.remove(message.from_user.id)
        await message.answer("Спасибо за ваш отзыв!")

    elif message.text == "Задать вопрос":
        questions.add(message.from_user.id)
        await message.answer("Задайте свой вопрос")

    elif message.from_user.id in questions:
        save_question(message.from_user.username or "No username", message.text)
        questions.remove(message.from_user.id)
        await message.answer("Спасибо за ваш вопрос! Ждите ответ в нашем телеграм сообществе.")
async def main():
    bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    await dp.start_polling(bot)

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, stream=sys.stdout)
    asyncio.run(main())
